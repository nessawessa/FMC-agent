"""
Copilot Context: FM&C Automation Agent Development Guide

This file provides context and guidance for AI assistants working on the FM&C Automation Agent project.
It includes architectural patterns, coding standards, and development workflows to maintain consistency.

## Project Overview

The FM&C Automation Agent is a Python CLI tool that:
1. Validates Excel workbooks containing FM&C (Failure Mode & Criticality) data
2. Parses operation sheets into executable commands
3. Executes or simulates RV&S CLI commands
4. Maintains audit trails in Change Log sheets

## Architecture Patterns

### Package Structure
```
src/fmc_agent/
├── __init__.py          # Version export and package metadata
├── cli.py              # Typer-based CLI interface (entry point)
├── logging_setup.py    # Centralized logging configuration
├── schemas.py          # Sheet schema definitions and validation rules
├── excel_validator.py  # Workbook structure validation
├── operations.py       # Operation parsing, command generation, execution
├── mock_data.py        # Sample data for testing and examples
└── utils/
    └── io.py           # File I/O utilities and helpers
```

### Design Principles

1. **Separation of Concerns**: Each module has a single, well-defined responsibility
2. **Configuration Over Convention**: Schema-driven validation and command generation
3. **Fail-Safe Operations**: Extensive validation before any mutations
4. **Audit Trail**: All operations logged to Change Log sheet
5. **Dry-Run Support**: Simulation mode for safe testing

### Key Classes and Responsibilities

#### ExcelValidator (excel_validator.py)
- Validates workbook structure (sheets, columns)
- Reports missing or malformed elements
- Provides summary statistics
- Thread-safe, read-only operations

#### OperationParser (operations.py)
- Parses Excel sheets into Operation objects
- Applies business rules (skip processed rows)
- Generates RV&S CLI commands from templates
- Handles missing/optional data gracefully

#### CommandExecutor (operations.py)
- Executes RV&S CLI commands via subprocess
- Supports dry-run simulation
- Extracts IDs from command output using heuristics
- Provides timeout and error handling

#### ChangeLogManager (operations.py)
- Appends audit entries to Change Log sheet
- Handles Change Log sheet creation if missing
- Truncates long outputs for readability
- Thread-safe Excel write operations

## Coding Standards

### Python Style
- Follow PEP 8 with 100-character line limit
- Use type hints for all public functions
- Prefer composition over inheritance
- Use dataclasses for data containers
- Document all public APIs with docstrings

### Error Handling
- Use specific exception types where possible
- Log errors with appropriate levels (ERROR, WARNING, INFO, DEBUG)
- Provide actionable error messages to users
- Graceful degradation (continue processing when possible)

### Testing Patterns
- Use pytest for all tests
- Mock external dependencies (subprocess, file I/O)
- Test both success and failure paths
- Include integration tests with sample data

### Logging Strategy
- Use structured logging with consistent format
- Include context (sheet names, row numbers, operation types)
- Separate user-facing output (Rich console) from debug logs
- Configurable log levels via CLI options

## Development Workflows

### Adding New Operation Types

1. **Update schemas.py**: Add new sheet schema to SHEET_SCHEMAS
2. **Update operations.py**: Add command template to COMMAND_TEMPLATES
3. **Update mock_data.py**: Add sample data for testing
4. **Add tests**: Create test cases for the new operation type
5. **Update documentation**: Add operation to FMC_TEMPLATE_STRUCTURE.md

### Adding New Validation Rules

1. **Extend ExcelValidator**: Add new validation method
2. **Update schemas.py**: Define new validation constraints
3. **Add tests**: Test both valid and invalid cases
4. **Update error messages**: Provide clear, actionable feedback

### Enhancing CLI Interface

1. **Use Typer patterns**: Leverage existing option/argument patterns
2. **Rich formatting**: Use Rich for user-friendly output
3. **Error handling**: Exit with appropriate codes (0=success, 1=error)
4. **Documentation**: Update CLI help text and README

## Common Patterns

### File Path Handling
```python
from pathlib import Path

def process_file(file_path: Path) -> None:
    if not file_path.exists():
        raise FileNotFoundError(f"File not found: {file_path}")
    
    # Use pathlib methods for cross-platform compatibility
    backup_path = file_path.with_suffix(f".backup{file_path.suffix}")
```

### Excel Operations
```python
import pandas as pd
from openpyxl import load_workbook

# Read-only operations (use pandas for data processing)
df = pd.read_excel(file_path, sheet_name="Sheet Name")

# Write operations (use openpyxl for precise control)
workbook = load_workbook(file_path)
worksheet = workbook["Sheet Name"]
worksheet.cell(row=1, column=1, value="New Value")
workbook.save(file_path)
workbook.close()
```

### Command Execution
```python
import subprocess
from typing import Tuple

def execute_command(command: str) -> Tuple[bool, str]:
    try:
        result = subprocess.run(
            command, shell=True, capture_output=True, 
            text=True, timeout=300
        )
        return result.returncode == 0, result.stdout or result.stderr
    except subprocess.TimeoutExpired:
        return False, "Command timed out"
    except Exception as e:
        return False, f"Execution failed: {e}"
```

### Rich CLI Output
```python
from rich.console import Console
from rich.table import Table

console = Console()

# Status messages with color coding
console.print("[green]✓ Success[/green]")
console.print("[red]✗ Failed[/red]")
console.print("[yellow]⚠ Warning[/yellow]")

# Structured data display
table = Table(show_header=True, header_style="bold magenta")
table.add_column("Column 1")
table.add_column("Column 2")
table.add_row("Value 1", "Value 2")
console.print(table)
```

## Extension Points

### Plugin Architecture (Future)
- Define abstract base classes for operations
- Use entry points for plugin discovery
- Maintain backward compatibility with core operations

### Custom Validators (Future)
- Extensible validation framework
- Industry-specific validation rules
- Custom error message formatting

### Output Formatters (Future)
- JSON output for programmatic consumption
- Custom report templates
- Integration with external reporting systems

## Testing Guidelines

### Unit Tests
- Test individual functions and methods
- Mock external dependencies (files, subprocess)
- Use parametrized tests for multiple input scenarios
- Aim for >90% code coverage

### Integration Tests
- Test complete workflows end-to-end
- Use sample Excel files with known data
- Verify Change Log entries are created correctly
- Test CLI commands with various options

### Performance Tests
- Test with large Excel files (1000+ rows)
- Measure memory usage and execution time
- Identify bottlenecks in Excel processing
- Validate timeout handling

## Common Debugging Scenarios

### Excel File Issues
- File locked by another application
- Corrupted Excel files
- Missing sheets or columns
- Encoding issues with special characters

### RV&S CLI Issues
- Command not found in PATH
- Authentication/permission errors
- Network connectivity problems
- Unexpected output format changes

### Performance Issues
- Large Excel files consuming excessive memory
- Slow Excel read/write operations
- Subprocess timeout in high-latency environments
- Change Log sheet becoming very large

## Security Considerations

### Input Validation
- Sanitize all Excel cell content before using in commands
- Validate file paths to prevent directory traversal
- Limit subprocess command construction to prevent injection

### Credential Handling
- Never log credentials or sensitive data
- Use environment variables for authentication
- Implement credential rotation support

### Audit Trail
- Log all operations with timestamps
- Include user context in audit entries
- Maintain immutable audit records

This context should help maintain consistency and quality as the project evolves.
"""