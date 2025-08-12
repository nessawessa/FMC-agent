"""Operation parsing and command generation for RV&S CLI."""

import logging
import re
import subprocess
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


@dataclass
class Operation:
    """Represents a single operation to be executed."""

    operation_type: str
    row_data: Dict[str, str]
    command: str
    sheet_name: str
    row_index: int

    def __str__(self) -> str:
        return f"{self.operation_type} - Row {self.row_index}: {self.command[:100]}..."


class OperationParser:
    """Parses operations from Excel sheets and generates RV&S commands."""

    # Command templates for different operation types
    COMMAND_TEMPLATES = {
        "Create Fail Modes": (
            "im createissue --type='Fail Mode' --field='Functional System ID'='{Functional System ID}' "
            "--field='Name'='{Fail Mode Name}' --field='Description'='{Fail Mode Description}'"
        ),
        "Create Causes": (
            "im createissue --type='Cause' --field='Fail Mode ID'='{Fail Mode ID}' "
            "--field='Name'='{Cause Name}' --field='Description'='{Cause Description}'"
        ),
        "Create Controls": (
            "im createissue --type='Control' --field='Control Type'='{Control Type}' "
            "--field='Name'='{Control Name}' --field='Description'='{Control Description}'"
        ),
        "Create Control Causes": (
            "im createrelationship --type='Control-Cause' "
            "--field='Control ID'='{Control ID}' --field='Cause ID'='{Cause ID}'"
        ),
    }

    def __init__(self, file_path: Path):
        """Initialize parser with workbook file path.
        
        Args:
            file_path: Path to the Excel workbook file.
        """
        self.file_path = file_path

    def parse_operations(self, operation_types: Optional[List[str]] = None) -> List[Operation]:
        """Parse operations from specified sheets.
        
        Args:
            operation_types: List of operation types to parse. If None, parse all.
            
        Returns:
            List of Operation objects.
        """
        operations = []

        if operation_types is None:
            operation_types = list(self.COMMAND_TEMPLATES.keys())

        for op_type in operation_types:
            if op_type not in self.COMMAND_TEMPLATES:
                logger.warning(f"Unknown operation type: {op_type}")
                continue

            try:
                sheet_operations = self._parse_sheet(op_type)
                operations.extend(sheet_operations)
                logger.info(f"Parsed {len(sheet_operations)} operations from '{op_type}' sheet")
            except Exception as e:
                logger.error(f"Failed to parse operations from '{op_type}' sheet: {e}")

        return operations

    def _parse_sheet(self, sheet_name: str) -> List[Operation]:
        """Parse operations from a single sheet.
        
        Args:
            sheet_name: Name of the sheet to parse.
            
        Returns:
            List of Operation objects from the sheet.
        """
        operations = []

        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name)
        except Exception as e:
            logger.error(f"Failed to read sheet '{sheet_name}': {e}")
            return operations

        template = self.COMMAND_TEMPLATES[sheet_name]

        for idx, row in df.iterrows():
            # Skip rows where Agent Status indicates already processed
            agent_status = str(row.get("Agent Status", "")).strip().lower()
            if agent_status in ["completed", "success", "done"]:
                logger.debug(f"Skipping row {idx + 2} - already processed")
                continue

            # Convert row to dict, handling NaN values
            row_data = {}
            for col, value in row.items():
                if pd.isna(value):
                    row_data[col] = ""
                else:
                    row_data[col] = str(value).strip()

            # Generate command from template
            try:
                command = template.format(**row_data)
                operation = Operation(
                    operation_type=sheet_name,
                    row_data=row_data,
                    command=command,
                    sheet_name=sheet_name,
                    row_index=idx + 2,  # +2 because pandas is 0-indexed and Excel has header
                )
                operations.append(operation)
            except KeyError as e:
                logger.warning(f"Missing required field for row {idx + 2} in '{sheet_name}': {e}")

        return operations


class CommandExecutor:
    """Executes RV&S CLI commands and captures output."""

    def __init__(self, dry_run: bool = False):
        """Initialize executor.
        
        Args:
            dry_run: If True, simulate command execution without actually running them.
        """
        self.dry_run = dry_run

    def execute_operation(self, operation: Operation) -> Tuple[bool, str, Optional[str]]:
        """Execute a single operation.
        
        Args:
            operation: Operation to execute.
            
        Returns:
            Tuple of (success, output, extracted_id)
        """
        if self.dry_run:
            return self._simulate_execution(operation)
        return self._execute_command(operation)

    def _simulate_execution(self, operation: Operation) -> Tuple[bool, str, Optional[str]]:
        """Simulate command execution for dry-run mode.
        
        Args:
            operation: Operation to simulate.
            
        Returns:
            Tuple of (success, simulated_output, dummy_id)
        """
        # Generate a dummy ID for simulation
        dummy_id = f"SIM-{operation.operation_type.replace(' ', '')}-{operation.row_index:04d}"

        output = (
            f"[DRY-RUN] Would execute: {operation.command}\n"
            f"[DRY-RUN] Simulated success - Generated ID: {dummy_id}"
        )

        logger.info(f"Simulated execution of {operation.operation_type} - Row {operation.row_index}")
        return True, output, dummy_id

    def _execute_command(self, operation: Operation) -> Tuple[bool, str, Optional[str]]:
        """Execute the actual RV&S CLI command.
        
        Args:
            operation: Operation to execute.
            
        Returns:
            Tuple of (success, command_output, extracted_id)
        """
        try:
            logger.info(f"Executing: {operation.command}")
            result = subprocess.run(
                operation.command,
                check=False, shell=True,
                capture_output=True,
                text=True,
                timeout=300,  # 5 minute timeout
            )

            success = result.returncode == 0
            output = result.stdout if success else result.stderr

            # Extract ID from output using heuristic pattern matching
            extracted_id = self._extract_id_from_output(output)

            if success:
                logger.info(f"Command succeeded - ID: {extracted_id}")
            else:
                logger.error(f"Command failed with return code {result.returncode}")

            return success, output, extracted_id

        except subprocess.TimeoutExpired:
            error_msg = f"Command timed out after 300 seconds: {operation.command}"
            logger.error(error_msg)
            return False, error_msg, None
        except Exception as e:
            error_msg = f"Failed to execute command: {e}"
            logger.error(error_msg)
            return False, error_msg, None

    def _extract_id_from_output(self, output: str) -> Optional[str]:
        """Extract ID from command output using heuristic patterns.
        
        Args:
            output: Command output text.
            
        Returns:
            Extracted ID if found, None otherwise.
        """
        # Common patterns for ID extraction
        patterns = [
            r"ID:\s*(\d+)",
            r"Created.*?ID\s*(\d+)",
            r"Issue\s*(\d+)\s*created",
            r"#(\d+)",
            r"\b(\d{4,})\b",  # Generic 4+ digit number
        ]

        for pattern in patterns:
            match = re.search(pattern, output, re.IGNORECASE)
            if match:
                return match.group(1)

        return None


class ChangeLogManager:
    """Manages Change Log sheet updates."""

    def __init__(self, file_path: Path):
        """Initialize change log manager.
        
        Args:
            file_path: Path to the Excel workbook file.
        """
        self.file_path = file_path

    def append_entry(self, operation: Operation, success: bool, output: str,
                    extracted_id: Optional[str] = None) -> None:
        """Append an entry to the Change Log sheet.
        
        Args:
            operation: The operation that was executed.
            success: Whether the operation succeeded.
            output: Command output.
            extracted_id: Extracted ID from the output, if any.
        """
        try:
            # Truncate output to prevent excessive cell content
            truncated_output = output[:2000]
            if len(output) > 2000:
                truncated_output += "... [truncated]"

            entry_data = {
                "Timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Operation": f"{operation.operation_type} - Row {operation.row_index}",
                "Status": "Success" if success else "Failed",
                "Details": f"ID: {extracted_id}" if extracted_id else "No ID extracted",
                "CLI Output": truncated_output,
            }

            # Load workbook in read-write mode
            workbook = load_workbook(self.file_path)

            # Create Change Log sheet if it doesn't exist
            if "Change Log" not in workbook.sheetnames:
                ws = workbook.create_sheet("Change Log")
                # Add headers
                headers = ["Timestamp", "Operation", "Status", "Details", "CLI Output"]
                for col, header in enumerate(headers, 1):
                    ws.cell(row=1, column=col, value=header)
            else:
                ws = workbook["Change Log"]

            # Find next empty row
            next_row = ws.max_row + 1

            # Add the entry
            for col, (key, value) in enumerate(entry_data.items(), 1):
                ws.cell(row=next_row, column=col, value=value)

            # Save the workbook
            workbook.save(self.file_path)
            workbook.close()

            logger.info(f"Added Change Log entry for {operation.operation_type} - Row {operation.row_index}")

        except Exception as e:
            logger.error(f"Failed to append Change Log entry: {e}")

    def get_recent_entries(self, count: int = 10) -> List[Dict[str, str]]:
        """Get recent Change Log entries.
        
        Args:
            count: Number of recent entries to retrieve.
            
        Returns:
            List of recent Change Log entries.
        """
        try:
            df = pd.read_excel(self.file_path, sheet_name="Change Log")
            recent_df = df.tail(count)
            return recent_df.to_dict("records")
        except Exception as e:
            logger.error(f"Failed to read Change Log entries: {e}")
            return []
