"""Excel workbook validation functionality."""

import logging
from pathlib import Path
from typing import Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook

from .schemas import SHEET_SCHEMAS

logger = logging.getLogger(__name__)


class ExcelValidator:
    """Validates FM&C Modification Template workbooks."""

    def __init__(self, file_path: Path):
        """Initialize validator with workbook file path.
        
        Args:
            file_path: Path to the Excel workbook file.
        """
        self.file_path = file_path
        self.workbook = None
        self.validation_errors: List[str] = []

    def validate(self) -> Tuple[bool, List[str]]:
        """Validate the workbook structure and content.
        
        Returns:
            Tuple of (is_valid, list_of_errors)
        """
        self.validation_errors = []

        try:
            self.workbook = load_workbook(self.file_path, read_only=True)
        except Exception as e:
            self.validation_errors.append(f"Failed to load workbook: {e}")
            return False, self.validation_errors

        # Validate sheet presence and structure
        self._validate_sheets()

        # Validate column schemas
        self._validate_columns()

        # TODO: Add cell-level validation (placeholder for future enhancement)
        # self._validate_cell_content()

        is_valid = len(self.validation_errors) == 0
        logger.info(f"Validation completed. Valid: {is_valid}, Errors: {len(self.validation_errors)}")

        return is_valid, self.validation_errors

    def _validate_sheets(self) -> None:
        """Validate that required sheets are present."""
        available_sheets = set(self.workbook.sheetnames)
        required_sheets = set(SHEET_SCHEMAS.keys())

        missing_sheets = required_sheets - available_sheets
        if missing_sheets:
            self.validation_errors.extend([
                f"Missing required sheet: '{sheet}'" for sheet in missing_sheets
            ])

        logger.debug(f"Available sheets: {available_sheets}")
        logger.debug(f"Required sheets: {required_sheets}")

    def _validate_columns(self) -> None:
        """Validate column schemas for each sheet."""
        for sheet_name, required_columns in SHEET_SCHEMAS.items():
            if sheet_name not in self.workbook.sheetnames:
                continue  # Already reported as missing sheet

            try:
                # Read just the header row to check columns
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, nrows=0)
                available_columns = set(df.columns)
                required_columns_set = set(required_columns)

                missing_columns = required_columns_set - available_columns
                if missing_columns:
                    self.validation_errors.extend([
                        f"Sheet '{sheet_name}' missing required column: '{col}'"
                        for col in missing_columns
                    ])

                logger.debug(f"Sheet '{sheet_name}' - Available: {available_columns}")
                logger.debug(f"Sheet '{sheet_name}' - Required: {required_columns_set}")

            except Exception as e:
                self.validation_errors.append(
                    f"Failed to validate columns in sheet '{sheet_name}': {e}",
                )

    def _validate_cell_content(self) -> None:
        """Validate cell content (placeholder for future implementation)."""
        # TODO: Implement cell-level validation
        # - Check for empty required fields
        # - Validate ID formats
        # - Check for duplicate IDs within sheets
        # - Validate cross-sheet references

    def get_sheet_summary(self) -> Dict[str, Dict[str, int]]:
        """Get summary information about each sheet.
        
        Returns:
            Dictionary with sheet names as keys and summary stats as values.
        """
        summary = {}

        # Load workbook if not already loaded
        if self.workbook is None:
            try:
                self.workbook = load_workbook(self.file_path, read_only=True)
            except Exception as e:
                # Return error summary for all sheets if workbook can't be loaded
                for sheet_name in SHEET_SCHEMAS.keys():
                    summary[sheet_name] = {
                        "status": "error",
                        "error": f"Failed to load workbook: {e}",
                        "rows": 0,
                    }
                return summary

        for sheet_name in SHEET_SCHEMAS.keys():
            if sheet_name not in self.workbook.sheetnames:
                summary[sheet_name] = {"status": "missing", "rows": 0}
                continue

            try:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name)
                summary[sheet_name] = {
                    "status": "present",
                    "rows": len(df),
                    "columns": len(df.columns),
                }
            except Exception as e:
                summary[sheet_name] = {
                    "status": "error",
                    "error": str(e),
                    "rows": 0,
                }

        return summary
